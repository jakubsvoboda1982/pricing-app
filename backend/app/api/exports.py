from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import Product, Price, CatalogProduct, CompetitorProductPrice
from app.models.recommendation import PriceRecommendation
from app.middleware.auth import verify_token
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import csv
from datetime import datetime
from typing import Optional
from decimal import Decimal

router = APIRouter(prefix="/api/export", tags=["export"])

# ── Field definitions grouped by category ────────────────────────────────────

FIELD_GROUPS = [
    {
        "label": "Identifikace produktu",
        "color": "1E3A5F",
        "fields": [
            ("name",         "Název produktu",      True),
            ("sku",          "SKU",                 True),
            ("product_code", "PRODUCTNO",           True),
            ("ean",          "EAN",                 True),
            ("manufacturer", "Výrobce",             True),
            ("category",     "Kategorie",           True),
            ("description",  "Popis",               False),
            ("url_reference","URL e-shop (CZ)",     False),
            ("id",           "Product ID",          False),
        ],
    },
    {
        "label": "Ceny – vlastní",
        "color": "1B6B3A",
        "fields": [
            ("current_price_czk",          "Aktuální cena CZK",        True),
            ("current_price_eur",          "Aktuální cena EUR (SK)",    True),
            ("old_price",                  "Předchozí cena",            False),
            ("purchase_price_without_vat", "Nákupní cena bez DPH",     True),
            ("purchase_price_with_vat",    "Nákupní cena s DPH",       True),
            ("min_price",                  "Min. cena s DPH",           True),
            ("margin_czk",                 "Marže CZ (%)",              True),
            ("margin_sk",                  "Marže SK (%)",              False),
        ],
    },
    {
        "label": "Sklad & pozice",
        "color": "5B2C8D",
        "fields": [
            ("stock_quantity",  "Skladem (ks)",        True),
            ("hero_score",      "Hero skóre",          True),
            ("market",          "Primární trh",        False),
        ],
    },
    {
        "label": "Konkurence",
        "color": "9B3A1A",
        "fields": [
            ("lowest_competitor_price", "Nejnižší konkurent",     True),
            ("competitors_count",       "Počet sledov. URL",      True),
            ("price_vs_competition",    "Naše cena vs min. konk.", False),
        ],
    },
    {
        "label": "Doporučení cen",
        "color": "0E4C7A",
        "fields": [
            ("recommended_price",         "Doporučená cena (s DPH)", True),
            ("recommended_price_source",  "Zdroj doporučení",        True),
            ("rec_margin_change",         "Změna marže (%)",         False),
            ("rec_revenue_impact",        "Dopad na tržby (%)",      False),
            ("rec_status",                "Stav doporučení",         False),
        ],
    },
    {
        "label": "Systémové",
        "color": "555555",
        "fields": [
            ("created_at", "Vytvořeno",  False),
            ("updated_at", "Upraveno",   False),
        ],
    },
]

# Flat structures derived from groups
ALL_FIELDS = [fid for g in FIELD_GROUPS for fid, _, _ in g["fields"]]
FIELD_LABELS = {fid: label for g in FIELD_GROUPS for fid, label, _ in g["fields"]}
FIELD_DEFAULT = {fid: default for g in FIELD_GROUPS for fid, _, default in g["fields"]}
FIELD_GROUP_NAME = {fid: g["label"] for g in FIELD_GROUPS for fid, _, _ in g["fields"]}
FIELD_GROUP_COLOR = {fid: g["color"] for g in FIELD_GROUPS for fid, _, _ in g["fields"]}


# ── Data loaders (batch) ──────────────────────────────────────────────────────

def _get_all_prices_map(db: Session, product_ids: list) -> dict:
    """Per product: {market: Price} — latest price per market."""
    if not product_ids:
        return {}
    prices = (
        db.query(Price)
        .filter(Price.product_id.in_(product_ids))
        .order_by(Price.product_id, Price.market, Price.changed_at.desc())
        .all()
    )
    # {product_id: {market: Price}}
    result: dict = {}
    for p in prices:
        pid = str(p.product_id)
        mkt = p.market or "CZ"
        if pid not in result:
            result[pid] = {}
        if mkt not in result[pid]:
            result[pid][mkt] = p
    return result


def _get_catalog_map(db: Session, catalog_ids: list) -> dict:
    """Load catalog products by ID."""
    if not catalog_ids:
        return {}
    cats = db.query(CatalogProduct).filter(CatalogProduct.id.in_(catalog_ids)).all()
    return {str(c.id): c for c in cats}


def _get_competitor_map(db: Session, product_ids: list) -> dict:
    """Per product_id: list of competitor prices (only with value)."""
    if not product_ids:
        return {}
    cps = (
        db.query(CompetitorProductPrice)
        .filter(
            CompetitorProductPrice.product_id.in_(product_ids),
            CompetitorProductPrice.price.isnot(None),
        )
        .all()
    )
    result: dict = {}
    for cp in cps:
        pid = str(cp.product_id)
        if pid not in result:
            result[pid] = []
        result[pid].append(cp)
    return result


def _get_recommendation_map(db: Session, product_ids: list) -> dict:
    """Latest pending/approved recommendation per product."""
    if not product_ids:
        return {}
    recs = (
        db.query(PriceRecommendation)
        .filter(
            PriceRecommendation.product_id.in_(product_ids),
            PriceRecommendation.status.in_(["pending", "approved"]),
        )
        .order_by(PriceRecommendation.product_id, PriceRecommendation.created_at.desc())
        .all()
    )
    seen: dict = {}
    for r in recs:
        pid = str(r.product_id)
        if pid not in seen:
            seen[pid] = r
    return seen


def _get_products(
    db: Session,
    company_id,
    category: Optional[str],
    market: Optional[str],
    min_price: Optional[float],
    max_price: Optional[float],
    search: Optional[str],
) -> list:
    q = db.query(Product)
    if company_id:
        q = q.filter(Product.company_id == company_id)
    if category:
        q = q.filter(Product.category == category)
    if search:
        q = q.filter(Product.name.ilike(f"%{search}%"))
    products = q.all()

    if min_price is not None or max_price is not None or market:
        price_map = _get_all_prices_map(db, [p.id for p in products])
        filtered = []
        for p in products:
            mkt_prices = price_map.get(str(p.id), {})
            pr = mkt_prices.get(market) if market else (mkt_prices.get("CZ") or next(iter(mkt_prices.values()), None))
            if market and not pr:
                continue
            if pr and pr.current_price is not None:
                val = float(pr.current_price)
                if min_price is not None and val < min_price:
                    continue
                if max_price is not None and val > max_price:
                    continue
            filtered.append(p)
        return filtered
    return products


# ── Row builder ────────────────────────────────────────────────────────────────

def _build_row(product, prices_by_market: dict, catalog, competitor_list: list, rec, fields: list) -> list:
    """Build one export row. prices_by_market = {market: Price}."""
    cz_price = prices_by_market.get("CZ")
    sk_price = prices_by_market.get("SK")

    current_czk = float(cz_price.current_price) if cz_price and cz_price.current_price else None
    current_eur = float(sk_price.current_price) if sk_price and sk_price.current_price else None

    # Purchase price with VAT
    ppwv = None
    if product.purchase_price_without_vat:
        vat = float(getattr(product, 'purchase_vat_rate', None) or 12)
        ppwv = round(float(product.purchase_price_without_vat) * (1 + vat / 100), 2)

    # Margin CZ
    margin_czk = None
    if current_czk and ppwv and current_czk > 0:
        margin_czk = round((current_czk - ppwv) / current_czk * 100, 1)

    # Margin SK
    margin_sk = None
    RATE_EUR = 24.5
    if current_eur and ppwv and current_eur > 0:
        purchase_eur = ppwv / RATE_EUR
        margin_sk = round((current_eur - purchase_eur) / current_eur * 100, 1)

    # Competitor data (CZ only for comparison)
    czk_competitors = [cp for cp in competitor_list if (cp.currency or 'CZK') == 'CZK' and cp.price is not None]
    lowest_comp = float(min(cp.price for cp in czk_competitors)) if czk_competitors else None
    comp_count = len(competitor_list)

    price_vs_comp = None
    if current_czk and lowest_comp:
        price_vs_comp = round((current_czk - lowest_comp) / lowest_comp * 100, 1)

    # Old price
    old_price = float(cz_price.old_price) if cz_price and cz_price.old_price else None

    # Recommendation
    rec_price = float(rec.recommended_price_with_vat) if rec else None
    rec_source = None
    rec_margin_change = None
    rec_revenue_impact = None
    rec_status_label = None
    if rec:
        reasoning = rec.reasoning or {}
        src = reasoning.get('source', '')
        rec_source = 'Simulátor co-když' if src == 'simulator' else 'Doporučení cen'
        rec_margin_change = float(rec.margin_change_percent) if rec.margin_change_percent else None
        rec_revenue_impact = float(rec.expected_revenue_impact_percent) if rec.expected_revenue_impact_percent else None
        status_map = {'pending': 'Čeká na schválení', 'approved': 'Schváleno', 'applied': 'Aplikováno', 'rejected': 'Zamítnuto'}
        rec_status_label = status_map.get(rec.status, rec.status)

    # Stock
    stock = getattr(product, 'stock_quantity', None)
    divisor = max(1, getattr(product, 'stock_divisor', 1) or 1)
    stock_display = int(stock // divisor) if stock is not None else None

    values = {
        'id':                        str(product.id),
        'name':                      product.name or '',
        'sku':                       product.sku or '',
        'product_code':              getattr(product, 'product_code', None) or '',
        'ean':                       getattr(product, 'ean', None) or '',
        'manufacturer':              catalog.manufacturer if catalog else '',
        'category':                  product.category or '',
        'description':               product.description or '',
        'url_reference':             getattr(product, 'url_reference', None) or '',
        'current_price_czk':         current_czk if current_czk is not None else '',
        'current_price_eur':         current_eur if current_eur is not None else '',
        'old_price':                 old_price if old_price is not None else '',
        'purchase_price_without_vat': float(product.purchase_price_without_vat) if product.purchase_price_without_vat else '',
        'purchase_price_with_vat':   ppwv if ppwv is not None else '',
        'min_price':                 float(product.min_price) if product.min_price else '',
        'margin_czk':                margin_czk if margin_czk is not None else '',
        'margin_sk':                 margin_sk if margin_sk is not None else '',
        'stock_quantity':            stock_display if stock_display is not None else '',
        'hero_score':                getattr(product, 'hero_score', '') or '',
        'market':                    (cz_price.market if cz_price else None) or 'CZ',
        'lowest_competitor_price':   lowest_comp if lowest_comp is not None else '',
        'competitors_count':         comp_count if comp_count > 0 else '',
        'price_vs_competition':      (f"+{price_vs_comp:.1f} %" if price_vs_comp and price_vs_comp > 0
                                      else f"{price_vs_comp:.1f} %" if price_vs_comp is not None else ''),
        'recommended_price':         rec_price if rec_price is not None else '',
        'recommended_price_source':  rec_source or '',
        'rec_margin_change':         rec_margin_change if rec_margin_change is not None else '',
        'rec_revenue_impact':        rec_revenue_impact if rec_revenue_impact is not None else '',
        'rec_status':                rec_status_label or '',
        'created_at':                product.created_at.strftime('%Y-%m-%d') if getattr(product, 'created_at', None) else '',
        'updated_at':                product.updated_at.strftime('%Y-%m-%d') if getattr(product, 'updated_at', None) else '',
    }

    return [values.get(f, '') for f in fields]


# ── Parse & load common data ──────────────────────────────────────────────────

def _parse_and_load(
    fields: Optional[str],
    category: Optional[str],
    market: Optional[str],
    min_price: Optional[float],
    max_price: Optional[float],
    search: Optional[str],
    db: Session,
    company_id=None,
):
    selected = [f for f in (fields.split(',') if fields else ALL_FIELDS) if f in ALL_FIELDS]
    if not selected:
        selected = [f for f in ALL_FIELDS if FIELD_DEFAULT.get(f)]

    products = _get_products(db, company_id, category, market, min_price, max_price, search)
    product_ids = [p.id for p in products]

    prices_map   = _get_all_prices_map(db, product_ids)
    catalog_ids  = [p.catalog_product_id for p in products if p.catalog_product_id]
    catalog_map  = _get_catalog_map(db, catalog_ids)
    comp_map     = _get_competitor_map(db, product_ids)
    rec_map      = _get_recommendation_map(db, product_ids)

    return selected, products, prices_map, catalog_map, comp_map, rec_map


# ── XLSX ─────────────────────────────────────────────────────────────────────

GROUP_COLORS = {g["label"]: g["color"] for g in FIELD_GROUPS}

@router.get("/products/xlsx")
def export_products_xlsx(
    fields:    Optional[str]   = Query(None),
    category:  Optional[str]   = Query(None),
    market:    Optional[str]   = Query(None),
    min_price: Optional[float] = Query(None),
    max_price: Optional[float] = Query(None),
    search:    Optional[str]   = Query(None),
    token_payload: dict = Depends(verify_token),
    db: Session = Depends(get_db),
):
    from uuid import UUID
    from app.models import User
    try:
        user_id = UUID(token_payload.get("sub"))
        user = db.query(User).filter(User.id == user_id).first()
        company_id = user.company_id if user else None
    except Exception:
        company_id = None

    selected, products, prices_map, catalog_map, comp_map, rec_map = _parse_and_load(
        fields, category, market, min_price, max_price, search, db, company_id
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produkty"
    ws.freeze_panes = "B3"

    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, bottom=thin)

    # Row 1: group headers (merged cells)
    # Row 2: field labels
    # Row 3+: data

    col = 1
    group_spans: list[tuple[str, str, int, int]] = []  # (label, color, col_start, col_end)
    for g in FIELD_GROUPS:
        g_fields = [fid for fid, _, _ in g["fields"] if fid in selected]
        if not g_fields:
            continue
        start = col
        end = col + len(g_fields) - 1
        group_spans.append((g["label"], g["color"], start, end))
        col = end + 1

    # Write group header row
    for label, color, start, end in group_spans:
        cell = ws.cell(row=1, column=start, value=label)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    ws.row_dimensions[1].height = 22

    # Write field label row
    col = 1
    for g in FIELD_GROUPS:
        g_fields = [fid for fid, _, _ in g["fields"] if fid in selected]
        for fid in g_fields:
            cell = ws.cell(row=2, column=col, value=FIELD_LABELS.get(fid, fid))
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell.font = Font(bold=True, size=9, color="374151")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            col += 1
    ws.row_dimensions[2].height = 30

    # Data rows — reorder selected to match group order
    ordered_fields = [fid for g in FIELD_GROUPS for fid, _, _ in g["fields"] if fid in selected]
    data_fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    for row_idx, product in enumerate(products, start=3):
        pid = str(product.id)
        row_data = _build_row(
            product,
            prices_map.get(pid, {}),
            catalog_map.get(str(product.catalog_product_id)) if product.catalog_product_id else None,
            comp_map.get(pid, []),
            rec_map.get(pid),
            ordered_fields,
        )
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = data_fill_even
            # Numeric alignment
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Column widths
    width_hints = {
        'name': 32, 'description': 32, 'id': 38, 'url_reference': 28,
        'sku': 18, 'product_code': 14, 'ean': 18, 'manufacturer': 18,
        'category': 22, 'price_vs_competition': 20,
        'recommended_price_source': 22, 'rec_status': 20,
    }
    for col_idx, fid in enumerate(ordered_fields, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_hints.get(fid, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"products-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── CSV ──────────────────────────────────────────────────────────────────────

@router.get("/products/csv")
def export_products_csv(
    fields:    Optional[str]   = Query(None),
    category:  Optional[str]   = Query(None),
    market:    Optional[str]   = Query(None),
    min_price: Optional[float] = Query(None),
    max_price: Optional[float] = Query(None),
    search:    Optional[str]   = Query(None),
    token_payload: dict = Depends(verify_token),
    db: Session = Depends(get_db),
):
    from uuid import UUID
    from app.models import User
    try:
        user_id = UUID(token_payload.get("sub"))
        user = db.query(User).filter(User.id == user_id).first()
        company_id = user.company_id if user else None
    except Exception:
        company_id = None

    selected, products, prices_map, catalog_map, comp_map, rec_map = _parse_and_load(
        fields, category, market, min_price, max_price, search, db, company_id
    )
    ordered_fields = [fid for g in FIELD_GROUPS for fid, _, _ in g["fields"] if fid in selected]

    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_ALL)
    w.writerow([FIELD_LABELS.get(f, f) for f in ordered_fields])
    for product in products:
        pid = str(product.id)
        w.writerow(_build_row(
            product,
            prices_map.get(pid, {}),
            catalog_map.get(str(product.catalog_product_id)) if product.catalog_product_id else None,
            comp_map.get(pid, []),
            rec_map.get(pid),
            ordered_fields,
        ))

    content = buf.getvalue().encode('utf-8-sig')
    filename = f"products-{datetime.now().strftime('%Y-%m-%d')}.csv"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Meta ─────────────────────────────────────────────────────────────────────

@router.get("/products/meta")
def export_meta(
    token_payload: dict = Depends(verify_token),
    db: Session = Depends(get_db),
):
    from sqlalchemy import distinct
    from uuid import UUID
    from app.models import User
    try:
        user_id = UUID(token_payload.get("sub"))
        user = db.query(User).filter(User.id == user_id).first()
        company_id = user.company_id if user else None
    except Exception:
        company_id = None

    q = db.query(Product)
    if company_id:
        q = q.filter(Product.company_id == company_id)
    cats = [r[0] for r in q.with_entities(distinct(Product.category)).filter(Product.category.isnot(None)).all()]
    markets_q = db.query(distinct(Price.market)).all()
    total = q.count()

    return {
        "categories": sorted(cats),
        "markets": [r[0] for r in markets_q],
        "total": total,
        "field_groups": [
            {
                "label": g["label"],
                "color": "#" + g["color"],
                "fields": [
                    {"id": fid, "label": label, "default": default}
                    for fid, label, default in g["fields"]
                ],
            }
            for g in FIELD_GROUPS
        ],
    }
